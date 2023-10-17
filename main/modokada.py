"""
modulo con funciones que calculan la deformacion a partir del slip utilizando las funciones del modulo dtopotools
"""

from clawpack.geoclaw import dtopotools, topotools
import numpy as np
from multiprocessing import Pool, Process, cpu_count
import shapefile
import pandas as pd
from openpyxl import load_workbook

# A partir de la falla creada y del slip calculado para ella, se crea los objetos dtopo.fault y dtopo.subfault, con los que se calcula la deformacion

def okada_solucion(lons, lats, razon_aspecto, strike, dip, prof, rake, slip, largo_falla, resolucion = 1/30., tamano_buffer = 2., verbose = False ):

    """
    Calcula la deformacion provocada por la distribucion de slip generada resolviendo okada
    utilizando las herramientas de clawpack dtopotools. Se crea un objeto Falla 
    generando listas con los atributos de cada subfalla
    Entradas:
    lons: array con las longitudes de cada subfalla
    lats: array con las latitudes de cada subfalla
    razon_aspecto: razon de aspecto de la falla (largo/ancho) 
    strike: array con los strikes de cada subfalla
    dip: array con los dips de cada subfalla
    prof: array con la profundidad de cada subfalla en metros
    rake: array con los rakes de cada subfalla
    slip: array con el slip de cada subfalla generado con la metodologia de LeVeque, 2016
    largo_falla: largo de la falla en metros
    resolucion: resolucion espacial con la que se calculara la deformacion (valor default: 1 minuto)
    tamano_buffer: tamano (en grados) de espacio "extra" en los bordes de la falla (valor default: 2 grados)
    verbose: booleano para mostrar por terminal el estado de la resolucion de okada. True: muestra mensaje False: no muestra mensaje (valor default: False)
    """

    # se crea el objeto falla
    Falla = dtopotools.Fault() # instancia Falla de la clase dtopotools.Fault()

    # parametros auxiliares para creacion de subfallas
    n_subfallas = int( np.size( lats ) ) # cantidad total de subfallas
    n_subfallas_filas = int( np.shape( lats )[0] ) # numero de filas (latitudes)
    n_subfallas_columnas = int( np.shape( lats )[1] ) # numero de columnas (longitudes)
    ancho_falla = largo_falla / razon_aspecto # ancho total de la falla
    largo_subfalla = largo_falla / n_subfallas_filas # largo de cada subfalla
    ancho_subfalla = ancho_falla / n_subfallas_columnas # ancho de cada subfalla

    # creacion de subfallas
    Falla.subfaults = [] # inicializacion lista con subfallas, se anadira a esta lista los atributos de cada subfalla
    for subfalla in range(int(n_subfallas)):
        SubFalla = dtopotools.SubFault() # iniciacion de instancia SubFalla de la clase dtopotools.SubFault. A este objeto se le anadiran los atributos que siguen para crear la falla "Falla"
        SubFalla.latitude = np.reshape( lats, ( np.size( lats ),) )[subfalla] # latitud de cada subfalla
        SubFalla.longitude = np.reshape( lons,( np.size( lons ),))[subfalla] # longitud de cada subfalla
        SubFalla.strike = np.reshape( strike, ( np.size( strike ),))[subfalla] # strike de cada subfalla
        SubFalla.rake = np.reshape( rake,( np.size( rake ),))[subfalla] # rake de cada subfalla
        SubFalla.depth = np.reshape( prof, ( np.size( prof ),))[subfalla] # profundidad de cada subfalla en metros
        SubFalla.dip = np.reshape( dip,( np.size( dip ),))[subfalla] # dip de cada subfalla
        SubFalla.slip = np.reshape( slip, ( np.size( slip ),))[subfalla] # slip de cada subfalla 
        SubFalla.length = (np.ones( np.size( slip )) * largo_subfalla)[subfalla] # largo de cada subfalla en metros
        SubFalla.width = (np.ones( np.size( slip )) * ancho_subfalla)[subfalla] # ancho de cada subfalla en metros
        SubFalla.coordinate_specification = 'centroid'
        Falla.subfaults = np.append(Falla.subfaults, SubFalla)

    # se crea la topografia  
    x,y = Falla.create_dtopo_xy( dx = resolucion, buffer_size = tamano_buffer ) 
    # se calcula la deformacion
    dtopo = Falla.create_dtopography(x,y,times=[1.], verbose = verbose ) # deformacion

    return dtopo


# crea el objeto falla de dtopotools a partir de los atributos de cada subfalla, para efectos de visualizacion y otros

def crea_falla_dtopo(lons, lats, razon_aspecto, strike, dip, prof, rake, slip, largo_falla):

    """
    Crea un objeto Falla 
    generando listas con los atributos de cada subfalla
    Entradas:
    lons: array con las longitudes de cada subfalla
    lats: array con las latitudes de cada subfalla
    razon_aspecto: razon de aspecto de la falla (largo/ancho) 
    strike: array con los strikes de cada subfalla
    dip: array con los dips de cada subfalla
    prof: array con la profundidad de cada subfalla en metros
    rake: array con los rakes de cada subfalla
    slip: array con el slip de cada subfalla generado con la metodologia de LeVeque, 2016
    largo_falla: largo de la falla en metros
    """

    # se crea el objeto falla
    Falla = dtopotools.Fault() # instancia Falla de la clase dtopotools.Fault()

    # parametros auxiliares para creacion de subfallas
    n_subfallas = int( np.size( lats ) ) # cantidad total de subfallas
    n_subfallas_filas = int( np.shape( lats )[0] ) # numero de filas (latitudes)
    n_subfallas_columnas = int( np.shape( lats )[1] ) # numero de columnas (longitudes)
    ancho_falla = largo_falla / razon_aspecto # ancho total de la falla
    largo_subfalla = largo_falla / n_subfallas_filas # largo de cada subfalla
    ancho_subfalla = ancho_falla / n_subfallas_columnas # ancho de cada subfalla

    # creacion de subfallas
    Falla.subfaults = [] # inicializacion lista con subfallas, se anadira a esta lista los atributos de cada subfalla
    for subfalla in range(int(n_subfallas)):
        SubFalla = dtopotools.SubFault() # iniciacion de instancia SubFalla de la clase dtopotools.SubFault. A este objeto se le anadiran los atributos que siguen para crear la falla "Falla"
        SubFalla.latitude = np.reshape( lats, ( np.size( lats ),) )[subfalla] # latitud de cada subfalla
        SubFalla.longitude = np.reshape( lons,( np.size( lons ),))[subfalla] # longitud de cada subfalla
        SubFalla.strike = np.reshape( strike, ( np.size( strike ),))[subfalla] # strike de cada subfalla
        SubFalla.rake = np.reshape( rake,( np.size( rake ),))[subfalla] # rake de cada subfalla
        SubFalla.depth = np.reshape( prof, ( np.size( prof ),))[subfalla] # profundidad de cada subfalla en metros
        SubFalla.dip = np.reshape( dip,( np.size( dip ),))[subfalla] # dip de cada subfalla
        SubFalla.slip = np.reshape( slip, ( np.size( slip ),))[subfalla] # slip de cada subfalla 
        SubFalla.length = (np.ones( np.size( slip )) * largo_subfalla)[subfalla] # largo de cada subfalla en metros
        SubFalla.width = (np.ones( np.size( slip )) * ancho_subfalla)[subfalla] # ancho de cada subfalla en metros
        SubFalla.coordinate_specification = 'centroid'
        Falla.subfaults = np.append(Falla.subfaults, SubFalla)

    return Falla


# calcula la magnitud de momento con las herramientas nativas de clawpack
def calcula_Mw( lons, lats, razon_aspecto, strike, dip, prof, rake, slip, largo_falla ):
    """
    con la funcion nativa de dtopotools .Mw() se calcula la magnitud de momento de la distribucion de slip
    Entradas:
    lons: array con las longitudes de cada subfalla
    lats: array con las latitudes de cada subfalla
    razon_aspecto: razon de aspecto de la falla (largo/ancho) 
    strike: array con los strikes de cada subfalla
    dip: array con los dips de cada subfalla
    prof: array con la profundidad de cada subfalla en metros
    rake: array con los rakes de cada subfalla
    slip: array con el slip de cada subfalla generado con la metodologia de LeVeque, 2016
    largo_falla: largo de la falla en metros
    """


    # se crea el objeto falla
    Falla = crea_falla_dtopo( lons, lats, razon_aspecto, strike, dip, prof, rake, slip, largo_falla )
    # se calcula su magnitud de momento
    Mw = Falla.Mw()
    return Mw

# funcion para guardar las salidas de deformacion

def guardar_okada( dtopo, nombre_archivo, tipo = 3 ):
    """
    Funcion para guardar las salidas del calculo de la deformacion del suelo oceanico
    a partir de la distribucion de slip
    Entradas:
    dtopo: deformacion oceanica calculada con el modelo de okada
    nombre_archivo: nombre del archivo de salida
    tipo: tipo de data a guardar, por defecto se utiliza dtopo_type = 3
    """

    # revisar si nombre_archivo es del tipo de string
    if not isinstance(nombre_archivo, basestring):
        nombre_archivo = str( nombre_archivo )
    else:
        nombre_archivo = nombre_archivo

    # se guarda el archivo 
    dtopo.write(nombre_archivo, dtopo_type = tipo)

# funcion para leer las salidas de deformacion

def leer_okada( nombre_archivo, directorio = None, tipo = 3 ):
    """
    Funcion para leer las salidas del calculo de la deformacion del suelo oceanico obtenidas
    con dtopo. Por defecto buscara en el cwd, pero se puede especificar un directorio si se
    desea leer un archivo ubicado en otro directorio.
    Entradas:
    nombre_archivo: nombre del archivo tt3 a leer
    directorio: por defecto se trabajara en el directorio actual, se puede especificar otro si se desea
    tipo: tipo de archivo dtopo a leer. Por defecto se utilizara en dtopo_type = 3
    """

    # se crea el objeto dtopo a llenarse con informacion de la deformacion leida
    dtopo = dtopotools.DTopography()

    # se crea el nombre del archivo a leerse, dependiendo del directorio desde donde se desee cargar
    if directorio == None:
        archivo = nombre_archivo
    elif isinstance( directorio, basestring ) and not nombre_archivo.startswith("/"):
        archivo = directorio + "/" + nombre_archivo
    elif isinstance( directorio, basestring ) and nombre_archivo.startswith("/"):
        archivo = directorio + nombre_archivo
    else:
        raise ValueError('Error en nombre de directorio o archivo')

    # se lee el archivo
    dtopo.read( archivo, dtopo_type = tipo )

    return dtopo

def cargar_deformacion(nombre_archivo, directorio = None, tipo = 3):
    """
    Funcion para cargar los valores del calculo de la deformacion del suelo oceanico obtenidas
    con dtopo. Por defecto buscara en el cwd, pero se puede especificar un directorio si se
    desea leer un archivo ubicado en otro directorio.
    Entradas:
    nombre_archivo: nombre del archivo tt3 a leer
    directorio: por defecto se trabajara en el directorio actual, se puede especificar otro si se desea
    tipo: tipo de archivo dtopo a leer. Por defecto se utilizara en dtopo_type = 3
    """

        # se crea el objeto dtopo a llenarse con informacion de la deformacion leida
    dtopo = dtopotools.DTopography()

    # se crea el nombre del archivo a leerse, dependiendo del directorio desde donde se desee cargar
    if directorio == None:
        archivo = nombre_archivo
    elif isinstance( directorio, basestring ) and not nombre_archivo.startswith("/"):
        archivo = directorio + "/" + nombre_archivo
    elif isinstance( directorio, basestring ) and nombre_archivo.startswith("/"):
        archivo = directorio + nombre_archivo
    else:
        raise ValueError('Error en nombre de directorio o archivo')

    # se lee el archivo
    dtopo.read( archivo, dtopo_type = tipo )

    # se extrae info de lons, lats y def
    X  = dtopo.X  # array con longitudes
    Y  = dtopo.Y  # array con latitudes
    dZ = dtopo.dZ # array con deformacion

    return X, Y, dZ 

# funcion para leer shapefiles con deformacion 

def leer_shape_deformacion( nombre_shape ):
    """
    Funcion para leer shapefiles con datos historicos de deformacion postsismica de terremotos en formato shp
    de acuerdo a los archivos entregados por cristian s
    Entradas: nombre de archivo shp. Especificar ruta relativa del archivo shp sin considerar el directorio madre de deformacion
    """

    # directorio donde se guardan datos de deformacion
    def_dir      = "../deformacion"
    # se asegura de que la ruta este bien definida
    if not nombre_shape.startswith("/"):
        nombre_shape = "/" + nombre_shape
    ruta_def_shp = def_dir + nombre_shape 
    
    # se crea el lector
    shp_def = shapefile.Reader(ruta_def_shp)
    # se asegura que el shapefile sea de tipo point (1) o multipoint (2)
    if not shp_def.shapeType == 1 or shp_def.shapeType == 8:
        raise ValueError('Error en formato de shapefile')
    

# funcion para leer archivos xls de deformacion otorgados por dr. patricio winkler

def leer_xls_observaciones_1960( nombre_xls ):
    """
    Funcion para leer los archivos xls de datos de observaciones del terremoto y tsunami de 1960
    otorgados por el dr. patricio winkler
    el formato de estos archivos es, por columna
    n sitio , nombre sitio, lat, lon, runup, profundidad de agua, altura de ola, inundacion, cambio de tierra, rango de marea, autor, comentarios
    Entradas: 
    nombre_xls: nombre del archivo con datos
    """


    # ruta relativa del archivo
    ruta_archivo   = "../deformacion/Registros_1960/"
    # nombre archivo
    nombre_archivo = "Registros-2019-02-17.xlsx"

    # se carga el workbook (contiene mas de un sheet el archivo)
    wb = load_workbook(ruta_archivo+nombre_archivo, read_only = True)
    # nos interesa solo el primer sheet, que contiene los datos importantes
    nombre_sheet = wb.sheetnames[0]